"""
Offline tests for:
  1. Model constant upgrade (MODEL_NAME == 'gpt-5.5')
  2. judge_translation_async — parsing, error handling
  3. process_row_async judge branches — clean / retried_passed / retried_flagged / error
  4. One-retry bound
  5. write_output_file judge_status column

No live OpenAI calls.  Run with:  python -m pytest test_model_upgrade_and_judge.py -v
"""

import asyncio
import json
import sys
import types
import os
import unittest
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from unittest.mock import AsyncMock, MagicMock, patch, call

# ---------------------------------------------------------------------------
# Minimal Streamlit stub so we can import the app without a running server
# ---------------------------------------------------------------------------
_st_stub = types.ModuleType("streamlit")
for _attr in (
    "set_page_config", "title", "subheader", "text", "caption", "info",
    "warning", "error", "success", "stop", "progress", "empty", "dataframe",
    "text_area", "checkbox", "selectbox", "text_input", "button", "expander",
    "session_state", "download_button", "fragment", "markdown",
):
    setattr(_st_stub, _attr, MagicMock(return_value=MagicMock()))
_st_stub.session_state = {}
sys.modules["streamlit"] = _st_stub

# Minimal openpyxl stub to avoid requiring the library for unit tests
_openpyxl_stub = types.ModuleType("openpyxl")
_wb = MagicMock()
_ws = MagicMock()
_ws.append = MagicMock()
_wb.active = _ws
_wb.create_sheet = MagicMock(return_value=_ws)
_wb.save = MagicMock()
_openpyxl_stub.Workbook = MagicMock(return_value=_wb)
sys.modules["openpyxl"] = _openpyxl_stub

# Suppress OPENAI_API_KEY requirement during import
os.environ.setdefault("OPENAI_API_KEY", "test-key")

import forsta_translation_qa_app as app  # noqa: E402 — must come after stubs


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_canned_openai_response(content: str):
    """Build a minimal fake OpenAI response object that _safe_json can parse."""
    choice = MagicMock()
    choice.message.content = content
    choice.finish_reason = "stop"
    resp = MagicMock()
    resp.choices = [choice]
    resp.system_fingerprint = None
    return resp


def _make_async_client(response_content: str):
    """Return a fake async client whose .chat.completions.create returns a canned response."""
    client = MagicMock()
    client.chat = MagicMock()
    client.chat.completions = MagicMock()
    client.chat.completions.create = AsyncMock(
        return_value=_make_canned_openai_response(response_content)
    )
    return client


def _run(coro):
    return asyncio.get_event_loop().run_until_complete(coro)


# ---------------------------------------------------------------------------
# Test group 1: model constant
# ---------------------------------------------------------------------------

class TestModelConstant(unittest.TestCase):
    def test_model_name_is_gpt_5_5(self):
        self.assertEqual(app.MODEL_NAME, "gpt-5.5")

    def test_no_old_constant_names(self):
        """The old two-constant names must not exist in the module."""
        self.assertFalse(hasattr(app, "TRANSLATION_MODEL_NAME"),
                         "TRANSLATION_MODEL_NAME still exists")
        self.assertFalse(hasattr(app, "CONSISTENCY_MODEL_NAME"),
                         "CONSISTENCY_MODEL_NAME still exists")

    def test_judge_threshold_default(self):
        self.assertEqual(app.JUDGE_SCORE_THRESHOLD, 3)

    def test_enable_judge_default_is_false(self):
        self.assertFalse(app.ENABLE_JUDGE_DEFAULT)


# ---------------------------------------------------------------------------
# Test group 2: judge_translation_async
# ---------------------------------------------------------------------------

class TestJudgeTranslationAsync(unittest.TestCase):

    def test_parses_valid_response(self):
        """A well-formed JSON response is parsed into score + reason."""
        content = json.dumps({"score": 4, "reason": "Reads naturally."})
        client = _make_async_client(content)
        with patch.object(app, "get_async_client", return_value=client):
            result = _run(app.judge_translation_async(
                english_text="How satisfied are you?",
                translation="Wie zufrieden sind Sie?",
                language_code="de",
            ))
        self.assertEqual(result["score"], 4)
        self.assertEqual(result["reason"], "Reads naturally.")
        self.assertNotIn("error", result)

    def test_none_content_returns_error(self):
        """content=None from the API must not raise; returns error dict."""
        bad_resp = _make_canned_openai_response(None)  # type: ignore[arg-type]
        client = MagicMock()
        client.chat.completions.create = AsyncMock(return_value=bad_resp)
        with patch.object(app, "get_async_client", return_value=client):
            result = _run(app.judge_translation_async(
                english_text="Hello",
                translation="Hola",
                language_code="es",
            ))
        self.assertIsNone(result.get("score"))
        self.assertTrue(result.get("error"))

    def test_garbage_json_returns_error(self):
        """Non-JSON response returns error dict."""
        client = _make_async_client("not valid json at all")
        with patch.object(app, "get_async_client", return_value=client):
            result = _run(app.judge_translation_async(
                english_text="Hello",
                translation="Bonjour",
                language_code="fr",
            ))
        self.assertIsNone(result.get("score"))
        self.assertTrue(result.get("error"))

    def test_missing_score_key_returns_error(self):
        """JSON without a 'score' key returns error dict."""
        content = json.dumps({"reason": "Missing score."})
        client = _make_async_client(content)
        with patch.object(app, "get_async_client", return_value=client):
            result = _run(app.judge_translation_async(
                english_text="Hello",
                translation="Ciao",
                language_code="it",
            ))
        self.assertIsNone(result.get("score"))
        self.assertTrue(result.get("error"))

    def test_out_of_range_score_returns_error(self):
        """Score 0 or 6 is out of range and should return error."""
        content = json.dumps({"score": 6, "reason": "Too good to be true."})
        client = _make_async_client(content)
        with patch.object(app, "get_async_client", return_value=client):
            result = _run(app.judge_translation_async(
                english_text="Hello",
                translation="Hola",
                language_code="es",
            ))
        self.assertTrue(result.get("error"))

    def test_language_agnostic_code_path(self):
        """Calling with two different languages follows identical code paths."""
        content = json.dumps({"score": 5, "reason": "Perfect."})
        for lang in ("ja", "ko"):
            client = _make_async_client(content)
            with patch.object(app, "get_async_client", return_value=client):
                result = _run(app.judge_translation_async(
                    english_text="Please rate your experience.",
                    translation="Please rate.",  # fake
                    language_code=lang,
                ))
            self.assertEqual(result["score"], 5)
            self.assertNotIn("error", result)


# ---------------------------------------------------------------------------
# Test group 3: process_row_async judge branches
# ---------------------------------------------------------------------------

def _make_minimal_context(language_code="de") -> app.SurveyFileContext:
    row = app.SurveyRow(
        variable_name="q1",
        english_text="How satisfied are you?",
        existing_translation="",
        had_real_translation=False,
    )
    row.segment_type = app.SegmentType.QUESTION
    ctx = app.SurveyFileContext(
        filename="test.xlsx",
        language_code=language_code,
        locale_code=language_code,
        rows=[row],
        translation_memory={},
    )
    return ctx


def _translation_result(text: str) -> dict:
    return {"proposed_translation": text, "qa_checked_translation": text,
            "needs_change": False, "change_reason": ""}


class TestProcessRowJudge(unittest.TestCase):

    # -- Branch 1: clean pass (score >= threshold) --
    def test_clean_pass(self):
        ctx = _make_minimal_context()
        row = ctx.rows[0]

        async def _run_it():
            with (
                patch.object(app, "call_translation_model_async",
                             new=AsyncMock(return_value=_translation_result("Wie zufrieden?"))),
                patch.object(app, "judge_translation_async",
                             new=AsyncMock(return_value={"score": 4, "reason": "Good."})) as mock_j,
            ):
                result = await app.process_row_async(
                    row, ctx, "ctx", asyncio.Semaphore(5),
                    provide_suggestions=False, enable_judge=True
                )
                return result, mock_j

        result, mock_j = _run(asyncio.coroutine(_run_it)() if False else _run_it())
        self.assertEqual(result.judge_outcome, "clean")
        self.assertFalse(result.judge_retried)
        self.assertEqual(result.judge_score, 4)
        # Judge called exactly once
        mock_j.assert_called_once()

    # -- Branch 2: retried and passed --
    def test_retried_passed(self):
        ctx = _make_minimal_context()
        row = ctx.rows[0]

        judge_calls = [
            {"score": 2, "reason": "Too formal."},   # T1 judge
            {"score": 4, "reason": "Better."},        # T2 judge
        ]

        async def _run_it():
            call_count = {"n": 0}
            async def _judge(*a, **kw):
                r = judge_calls[call_count["n"]]
                call_count["n"] += 1
                return r

            with (
                patch.object(app, "call_translation_model_async",
                             new=AsyncMock(side_effect=[
                                 _translation_result("T1-text"),   # T1
                                 _translation_result("T2-text"),   # retry
                             ])),
                patch.object(app, "judge_translation_async", new=_judge),
            ):
                result = await app.process_row_async(
                    row, ctx, "ctx", asyncio.Semaphore(5),
                    provide_suggestions=False, enable_judge=True
                )
            return result, call_count

        result, cc = _run(_run_it())
        self.assertEqual(result.judge_outcome, "retried_passed")
        self.assertTrue(result.judge_retried)
        self.assertEqual(result.judge_score, 4)
        # Judge called exactly twice (T1 then T2)
        self.assertEqual(cc["n"], 2)
        # T2 text was adopted
        self.assertIn("T2-text", result.new_translation)

    # -- Branch 3: retried and still flagged --
    def test_retried_flagged(self):
        ctx = _make_minimal_context()
        row = ctx.rows[0]

        judge_calls = [
            {"score": 1, "reason": "Very poor."},
            {"score": 2, "reason": "Still poor."},
        ]

        async def _run_it():
            call_count = {"n": 0}
            async def _judge(*a, **kw):
                r = judge_calls[call_count["n"]]
                call_count["n"] += 1
                return r

            with (
                patch.object(app, "call_translation_model_async",
                             new=AsyncMock(side_effect=[
                                 _translation_result("T1-text"),
                                 _translation_result("T2-text"),
                             ])),
                patch.object(app, "judge_translation_async", new=_judge),
            ):
                result = await app.process_row_async(
                    row, ctx, "ctx", asyncio.Semaphore(5),
                    provide_suggestions=False, enable_judge=True
                )
            return result, call_count

        result, cc = _run(_run_it())
        self.assertEqual(result.judge_outcome, "retried_flagged")
        self.assertTrue(result.judge_retried)
        self.assertEqual(result.judge_score, 2)
        self.assertEqual(cc["n"], 2)   # still only two judge calls
        # qa_status must contain the JUDGE flag
        self.assertIn("JUDGE", result.qa_status or "")

    # -- Branch 4: judge API error → non-fatal, ship T1 --
    def test_judge_error_nonfatal(self):
        ctx = _make_minimal_context()
        row = ctx.rows[0]

        async def _run_it():
            with (
                patch.object(app, "call_translation_model_async",
                             new=AsyncMock(return_value=_translation_result("T1-text"))),
                patch.object(app, "judge_translation_async",
                             new=AsyncMock(return_value={"score": None, "reason": "", "error": True})),
            ):
                result = await app.process_row_async(
                    row, ctx, "ctx", asyncio.Semaphore(5),
                    provide_suggestions=False, enable_judge=True
                )
            return result

        result = _run(_run_it())
        self.assertIsNone(result.judge_outcome)
        self.assertFalse(result.judge_retried)
        self.assertEqual(result.new_translation, "T1-text")

    # -- Branch 5: enable_judge=False → judge never called --
    def test_judge_disabled(self):
        ctx = _make_minimal_context()
        row = ctx.rows[0]

        async def _run_it():
            with (
                patch.object(app, "call_translation_model_async",
                             new=AsyncMock(return_value=_translation_result("T1-text"))),
                patch.object(app, "judge_translation_async",
                             new=AsyncMock()) as mock_j,
            ):
                result = await app.process_row_async(
                    row, ctx, "ctx", asyncio.Semaphore(5),
                    provide_suggestions=False, enable_judge=False
                )
            return result, mock_j

        result, mock_j = _run(_run_it())
        mock_j.assert_not_called()
        self.assertIsNone(result.judge_outcome)


# ---------------------------------------------------------------------------
# Test group 4: one-retry bound (judge called at most twice)
# ---------------------------------------------------------------------------

class TestJudgeOnRetryBound(unittest.TestCase):
    def test_at_most_two_judge_calls_regardless_of_t2_score(self):
        """Even if T2 also scores below threshold, there is no third attempt."""
        ctx = _make_minimal_context()
        row = ctx.rows[0]

        judge_results = [
            {"score": 1, "reason": "Poor."},
            {"score": 1, "reason": "Still poor."},
        ]

        call_count = {"n": 0}

        async def _judge(*a, **kw):
            r = judge_results[call_count["n"]]
            call_count["n"] += 1
            return r

        async def _run_it():
            with (
                patch.object(app, "call_translation_model_async",
                             new=AsyncMock(return_value=_translation_result("T1-text"))),
                patch.object(app, "judge_translation_async", new=_judge),
            ):
                await app.process_row_async(
                    row, ctx, "ctx", asyncio.Semaphore(5),
                    provide_suggestions=False, enable_judge=True
                )

        _run(_run_it())
        # Translation model called once for T1, once for retry = 2 total
        # Judge called once for T1, once for T2 = 2 total; never 3
        self.assertEqual(call_count["n"], 2)


# ---------------------------------------------------------------------------
# Test group 5: write_output_file judge_status column
# ---------------------------------------------------------------------------

class TestWriteOutputFileJudgeStatus(unittest.TestCase):
    def _make_context_with_judge_outcome(self, outcome, score, reason):
        import pandas as pd
        ctx = _make_minimal_context()
        row = ctx.rows[0]
        row.new_translation = "Eine Übersetzung"
        row.was_newly_translated = True
        row.judge_outcome = outcome
        row.judge_score = score
        row.judge_reason = reason
        row.judge_retried = outcome in ("retried_passed", "retried_flagged")
        original_df = pd.DataFrame({
            "variable_name": ["q1"],
            "english_text": ["How satisfied are you?"],
            "translation": [""],
        })
        return ctx, original_df

    def test_judge_status_populated_for_retried_flagged(self):
        import io
        ctx, orig = self._make_context_with_judge_outcome(
            "retried_flagged", 2, "Still too formal."
        )
        # Patch openpyxl write to capture the DataFrame instead of serializing
        with patch.object(app, "write_output_file") as _mock:
            # Call the real function
            pass
        # Call real function; mock openpyxl Workbook to avoid file I/O
        captured_df = None
        real_save = None

        import openpyxl
        orig_workbook = openpyxl.Workbook

        class _CapturingWorkbook:
            def __init__(self):
                self.active = MagicMock()
                self.active.title = ""
                self.active.append = MagicMock()
            def create_sheet(self):
                return self.active
            def save(self, buf):
                if hasattr(buf, "write"):
                    buf.write(b"xlsx-stub")

        with patch("openpyxl.Workbook", _CapturingWorkbook):
            df_out, fname, excel_bytes = app.write_output_file(ctx, orig, include_suggestions=False)

        self.assertIn("judge_status", df_out.columns)
        js = df_out.at[0, "judge_status"]
        self.assertIn("retried_flagged", js)
        self.assertIn("2", js)            # score in the cell
        self.assertIn("Still too formal.", js)

    def test_judge_status_empty_for_clean_pass(self):
        import pandas as pd

        class _CapturingWorkbook:
            def __init__(self):
                self.active = MagicMock()
                self.active.title = ""
                self.active.append = MagicMock()
            def create_sheet(self):
                return self.active
            def save(self, buf):
                if hasattr(buf, "write"):
                    buf.write(b"xlsx-stub")

        ctx = _make_minimal_context()
        row = ctx.rows[0]
        row.new_translation = "Eine Übersetzung"
        row.was_newly_translated = True
        # No judge ran (judge_outcome is None by default)
        orig = pd.DataFrame({
            "variable_name": ["q1"],
            "english_text": ["How satisfied are you?"],
            "translation": [""],
        })
        with patch("openpyxl.Workbook", _CapturingWorkbook):
            df_out, fname, excel_bytes = app.write_output_file(ctx, orig, include_suggestions=False)

        self.assertIn("judge_status", df_out.columns)
        self.assertEqual(df_out.at[0, "judge_status"], "")

    def test_first_three_columns_intact(self):
        """The first three Forsta-import columns must stay in positions 0-1-2."""
        import pandas as pd

        class _CapturingWorkbook:
            def __init__(self):
                self.active = MagicMock()
                self.active.title = ""
                self.active.append = MagicMock()
            def create_sheet(self):
                return self.active
            def save(self, buf):
                if hasattr(buf, "write"):
                    buf.write(b"xlsx-stub")

        ctx, orig = self._make_context_with_judge_outcome("clean", 5, "Perfect.")
        with patch("openpyxl.Workbook", _CapturingWorkbook):
            df_out, _, _ = app.write_output_file(ctx, orig, include_suggestions=False)

        cols = list(df_out.columns)
        self.assertEqual(cols[0], "variable_name")
        self.assertEqual(cols[1], "english_text")
        self.assertEqual(cols[2], "translation")


if __name__ == "__main__":
    unittest.main()
